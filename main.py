# ============================================================
# SISTEMA INTEGRADO DE CONTAS A PAGAR E FORNECEDORES
# ============================================================

import tkinter as tk
import locale
from sqlite3 import Cursor
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime  # <-- Adicionado para corrigir o erro de 'datetime'
from tkcalendar import DateEntry  # <-- Componente de Calendário

# Adicionado para corrigir o erro NameError: 'EXCEL_DISPONIVEL'
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False


# ============================================================
# BANCO DE DADOS UNIFICADO
# ============================================================

def conectar():
    """Conecta ao banco de dados unificado do sistema."""
    return sqlite3.connect("sistema_financeiro.db")


def criar_tabelas():
    """Cria as tabelas necessárias se elas não existirem."""
    conn = conectar()
    cursor = conn.cursor()

    # Tabela de fornecedores atualizada com os novos campos do print
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS fornecedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cnpj TEXT NOT NULL,
            telefone TEXT,
            email TEXT
        )
    """)

    # Tabela de contas a pagar relacionada com fornecedores
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            descricao TEXT NOT NULL,
            valor REAL NOT NULL,
            vencimento TEXT NOT NULL,
            status TEXT DEFAULT 'Pendente',
            fornecedor_id INTEGER,
            data_pagamento TEXT,
            FOREIGN KEY (fornecedor_id) REFERENCES fornecedores(id)
        )
    """)

    # Adiciona a coluna data_pagamento caso o banco já exista sem ela
    try:
        cursor.execute("ALTER TABLE contas ADD COLUMN data_pagamento TEXT")
    except Exception:
        pass  # Coluna já existe, ignora o erro

    conn.commit()
    conn.close()


# ============================================================
# JANELA: GERENCIAR FORNECEDORES
# ============================================================

class JanelaFornecedores(tk.Toplevel):
    def __init__(self, parent, callback_atualizar=None):
        super().__init__(parent)
        self.title("Gerenciar Fornecedores")
        self.geometry("550x500")
        self.resizable(False, False)
        self.grab_set()
        self.callback_atualizar = callback_atualizar
        self.criar_widgets()
        self.carregar_fornecedores()

    def criar_widgets(self):
        # Formulário de Cadastro
        frame_cad = tk.LabelFrame(self, text="Novo Fornecedor", padx=10, pady=10)
        frame_cad.pack(fill="x", padx=10, pady=10)

        tk.Label(frame_cad, text="Nome:").grid(row=0, column=0, sticky="w", pady=2)
        self.entry_nome = tk.Entry(frame_cad, width=45)
        self.entry_nome.grid(row=0, column=1, padx=5, pady=2)

        tk.Label(frame_cad, text="CNPJ:").grid(row=1, column=0, sticky="w", pady=2)
        self.entry_cnpj = tk.Entry(frame_cad, width=45)
        self.entry_cnpj.grid(row=1, column=1, padx=5, pady=2)

        tk.Label(frame_cad, text="Telefone:").grid(row=2, column=0, sticky="w", pady=2)
        self.entry_telefone = tk.Entry(frame_cad, width=45)
        self.entry_telefone.grid(row=2, column=1, padx=5, pady=2)

        tk.Label(frame_cad, text="Email:").grid(row=3, column=0, sticky="w", pady=2)
        self.entry_email = tk.Entry(frame_cad, width=45)
        self.entry_email.grid(row=3, column=1, padx=5, pady=2)

        tk.Button(frame_cad, text="Adicionar Fornecedor", command=self.adicionar, bg="#4CAF50", fg="white",
                  width=20).grid(row=4, column=0, columnspan=2, pady=10)

        # Visualização da Tabela
        frame_lista = tk.LabelFrame(self, text="Fornecedores Cadastrados", padx=10, pady=10)
        frame_lista.pack(fill="both", expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(frame_lista, columns=("Nome", "CNPJ"), show="headings", height=6)
        self.tree.heading("Nome", text="Nome")
        self.tree.heading("CNPJ", text="CNPJ")
        self.tree.column("Nome", width=300)
        self.tree.column("CNPJ", width=180)
        self.tree.pack(fill="both", expand=True)

        tk.Button(self, text="Excluir Selecionado", command=self.excluir, bg="#f44336", fg="white").pack(pady=5)

    def carregar_fornecedores(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nome, cnpj FROM fornecedores ORDER BY nome")
        self.fornecedores = cursor.fetchall()
        conn.close()

        for f in self.fornecedores:
            self.tree.insert("", tk.END, iid=f[0], values=(f[1], f[2]))

    def adicionar(self):
        nome = self.entry_nome.get().strip()
        cnpj = self.entry_cnpj.get().strip()
        telefone = self.entry_telefone.get().strip()
        email = self.entry_email.get().strip()

        if not nome or not cnpj:
            messagebox.showwarning("Aviso", "Nome e CNPJ são obrigatórios!", parent=self)
            return

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO fornecedores (nome, cnpj, telefone, email) VALUES (?, ?, ?, ?)",
            (nome, cnpj, telefone, email)
        )
        conn.commit()
        conn.close()

        self.entry_nome.delete(0, tk.END)
        self.entry_cnpj.delete(0, tk.END)
        self.entry_telefone.delete(0, tk.END)
        self.entry_email.delete(0, tk.END)

        self.carregar_fornecedores()
        if self.callback_atualizar:
            self.callback_atualizar()
        messagebox.showinfo("Sucesso", "Fornecedor cadastrado!", parent=self)

    def excluir(self):
        selecionado = self.tree.selection()
        if not id(selecionado):
            if not selecionado:
                messagebox.showwarning("Atenção", "Selecione um fornecedor para excluir.", parent=self)
                return

        fornecedor_id = selecionado[0]
        if messagebox.askyesno("Confirmar", "Deseja realmente excluir este fornecedor?", parent=self):
            conn = conectar()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM fornecedores WHERE id = ?", (fornecedor_id,))
            conn.commit()
            conn.close()
            self.carregar_fornecedores()
            if self.callback_atualizar:
                self.callback_atualizar()


# ============================================================
# JANELA: CADASTRAR / EDITAR CONTA
# ============================================================

class JanelaConta(tk.Toplevel):
    def __init__(self, parent, callback_atualizar, conta_id=None):
        super().__init__(parent)
        self.callback_atualizar = callback_atualizar
        self.conta_id = conta_id
        self.title("Editar Conta" if conta_id else "Nova Conta")
        self.resizable(False, False)
        self.grab_set()
        self.duplicatas = []  # lista de dicts: {valor, vencimento, widgets}
        self.criar_widgets()
        self.carregar_fornecedores()
        if conta_id:
            self.preencher_dados()

    # ------------------------------------------------------------------ layout

    def criar_widgets(self):
        # Frame com scroll para suportar muitas duplicatas
        container = tk.Frame(self)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        self.scroll_frame = tk.Frame(canvas, padx=20, pady=15)

        self.scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.canvas = canvas

        # Rolar com a roda do mouse
        def _roda(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _roda)

        frame = self.scroll_frame

        # --- Campos do boleto principal ---
        tk.Label(frame, text="Boleto Principal", font=("Arial", 10, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 6))

        tk.Label(frame, text="Valor (R$):").grid(row=1, column=0, sticky="w", pady=5)
        self.entry_valor = tk.Entry(frame, width=15)
        self.entry_valor.grid(row=1, column=1, sticky="w", pady=5)

        tk.Label(frame, text="Vencimento (DD/MM/AAAA):").grid(row=2, column=0, sticky="w", pady=5)
        self.entry_vencimento = tk.Entry(frame, width=15)
        self.entry_vencimento.grid(row=2, column=1, sticky="w", pady=5)

        tk.Label(frame, text="Fornecedor:").grid(row=3, column=0, sticky="w", pady=5)
        self.combo_fornecedor = ttk.Combobox(frame, width=27, state="normal")
        self.combo_fornecedor.grid(row=3, column=1, sticky="w", pady=5, columnspan=2)
        self.combo_fornecedor.bind("<KeyRelease>", self.filtrar_fornecedores)

        tk.Label(frame, text="Status:").grid(row=4, column=0, sticky="w", pady=5)
        self.combo_status = ttk.Combobox(frame, values=["Pendente", "Pago"], width=12, state="readonly")
        self.combo_status.set("Pendente")
        self.combo_status.grid(row=4, column=1, sticky="w", pady=5)

        # Botao "+" apenas no modo Nova Conta
        if not self.conta_id:
            btn_add = tk.Button(
                frame, text="  +  Adicionar duplicata  ",
                command=self._adicionar_duplicata,
                bg="#FF9800", fg="white",
                font=("Arial", 9, "bold"),
                relief="flat", cursor="hand2"
            )
            btn_add.grid(row=5, column=0, columnspan=3, sticky="w", pady=(10, 4))

            self.sep_dup = tk.Frame(frame, height=2, bg="#e0e0e0")
            self.sep_dup.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(0, 6))
            self.sep_dup.grid_remove()  # esconde até ter duplicatas

        # Linha onde as duplicatas serao inseridas dinamicamente (row 7+)
        self.prox_row_dup = 7

        # --- Botoes de salvar ---
        self.frame_btns = tk.Frame(frame)
        self.frame_btns.grid(row=self.prox_row_dup, column=0, columnspan=3, pady=12)
        self._renderizar_botoes()

        self._ajustar_tamanho()

    def _renderizar_botoes(self):
        for w in self.frame_btns.winfo_children():
            w.destroy()
        tk.Button(self.frame_btns, text="Salvar Conta", command=self.salvar,
                  bg="#2196F3", fg="white", width=16).pack(side="left", padx=5)
        if not self.conta_id:
            tk.Button(self.frame_btns, text="Salvar e Fechar", command=self.salvar_e_fechar,
                      bg="#4CAF50", fg="white", width=16).pack(side="left", padx=5)

    def _ajustar_tamanho(self):
        """Recalcula a altura da janela conforme o número de duplicatas."""
        base = 310 if self.conta_id else 260
        altura = base + len(self.duplicatas) * 54
        altura = min(altura, 680)
        self.geometry(f"520x{altura}")
        self.canvas.configure(width=480)

    # ---------------------------------------------------------------- duplicatas

    def _adicionar_duplicata(self):
        """Copia o ultimo vencimento +30 dias e cria uma linha editavel."""
        # Descobre o vencimento de referencia (ultimo da lista ou o principal)
        if self.duplicatas:
            ref_venc_str = self.duplicatas[-1]["venc_var"].get()
            ref_val_str  = self.duplicatas[-1]["val_var"].get()
        else:
            ref_venc_str = self.entry_vencimento.get().strip()
            ref_val_str  = self.entry_valor.get().strip()

        # Calcula +30 dias
        try:
            ref_dt = datetime.strptime(ref_venc_str, "%d/%m/%Y")
            import calendar
            # Avanca mes a mes para respeitar duplicatas mensais
            mes = ref_dt.month + 1
            ano = ref_dt.year + (1 if mes > 12 else 0)
            mes = mes if mes <= 12 else 1
            dia = min(ref_dt.day, calendar.monthrange(ano, mes)[1])
            novo_venc = ref_dt.replace(year=ano, month=mes, day=dia).strftime("%d/%m/%Y")
        except ValueError:
            novo_venc = ""

        # Linha da duplicata na grade
        idx = len(self.duplicatas)
        row = self.prox_row_dup + idx

        # Reposiciona o frame de botoes e o separador
        self.frame_btns.grid(row=row + 1, column=0, columnspan=3, pady=12)

        num_label = tk.Label(self.scroll_frame,
                             text=f"Duplicata {idx + 1}",
                             font=("Arial", 9, "bold"), fg="#555")
        num_label.grid(row=row, column=0, sticky="w", pady=(6, 2))

        val_var  = tk.StringVar(value=ref_val_str)
        venc_var = tk.StringVar(value=novo_venc)

        frame_linha = tk.Frame(self.scroll_frame)
        frame_linha.grid(row=row, column=1, columnspan=2, sticky="w", pady=(6, 2))

        tk.Label(frame_linha, text="Valor R$:").pack(side="left")
        e_val = tk.Entry(frame_linha, textvariable=val_var, width=11)
        e_val.pack(side="left", padx=(3, 10))

        tk.Label(frame_linha, text="Vencimento:").pack(side="left")
        e_venc = tk.Entry(frame_linha, textvariable=venc_var, width=12)
        e_venc.pack(side="left", padx=(3, 8))

        def remover(i=idx):
            self._remover_duplicata(i)

        btn_rem = tk.Button(frame_linha, text="✕", command=remover,
                            bg="#f44336", fg="white", width=2,
                            relief="flat", cursor="hand2")
        btn_rem.pack(side="left")

        self.duplicatas.append({
            "val_var":    val_var,
            "venc_var":   venc_var,
            "num_label":  num_label,
            "frame_linha": frame_linha,
        })

        # Mostra o separador visual
        if hasattr(self, "sep_dup"):
            self.sep_dup.grid()

        self._ajustar_tamanho()
        # Foca no campo de valor da nova duplicata para facilitar edicao
        e_val.focus()
        e_val.selection_range(0, tk.END)

    def _remover_duplicata(self, idx):
        d = self.duplicatas[idx]
        d["num_label"].destroy()
        d["frame_linha"].destroy()
        self.duplicatas.pop(idx)

        # Renumera os rotulos restantes
        for i, dup in enumerate(self.duplicatas):
            dup["num_label"].config(text=f"Duplicata {i + 1}")

        if not self.duplicatas and hasattr(self, "sep_dup"):
            self.sep_dup.grid_remove()

        self._ajustar_tamanho()

    # ---------------------------------------------------------------- fornecedores

    def carregar_fornecedores(self):
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nome FROM fornecedores ORDER BY nome")
        self.lista_fornecedores = cursor.fetchall()
        conn.close()
        self.nomes_fornecedores = ["(Nenhum)"] + [f[1] for f in self.lista_fornecedores]
        self.combo_fornecedor["values"] = self.nomes_fornecedores
        self.combo_fornecedor.set("(Nenhum)")

    def filtrar_fornecedores(self, event):
        texto = self.combo_fornecedor.get().strip().lower()
        if texto == "":
            self.combo_fornecedor["values"] = self.nomes_fornecedores
        else:
            self.combo_fornecedor["values"] = [n for n in self.nomes_fornecedores if texto in n.lower()]
        self.combo_fornecedor.event_generate("<Down>")

    def preencher_dados(self):
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("SELECT valor, vencimento, status, fornecedor_id FROM contas WHERE id = ?",
                       (self.conta_id,))
        conta = cursor.fetchone()
        conn.close()
        if conta:
            val, venc, status, f_id = conta
            self.entry_valor.insert(0, str(val))
            self.entry_vencimento.insert(0, venc)
            self.combo_status.set(status)
            if f_id:
                for f in self.lista_fornecedores:
                    if f[0] == f_id:
                        self.combo_fornecedor.set(f[1])
                        break

    # ---------------------------------------------------------------- salvar

    def _validar_e_obter_dados(self):
        val_str = self.entry_valor.get().strip()
        venc    = self.entry_vencimento.get().strip()
        status  = self.combo_status.get()
        nome    = self.combo_fornecedor.get().strip()

        if not val_str or not venc:
            messagebox.showwarning("Aviso", "Preencha o Valor e o Vencimento do boleto principal.", parent=self)
            return None
        try:
            val = float(val_str.replace(",", "."))
        except ValueError:
            messagebox.showerror("Erro", "Valor numerico invalido.", parent=self)
            return None

        f_id = None
        if nome and nome != "(Nenhum)":
            for f in self.lista_fornecedores:
                if f[1].lower() == nome.lower():
                    f_id = f[0]
                    break
            if f_id is None:
                messagebox.showwarning("Aviso",
                    "Fornecedor nao encontrado! Cadastre-o primeiro ou selecione um valido.",
                    parent=self)
                return None
        return val, venc, status, f_id

    def _gravar(self, fechar=False):
        dados = self._validar_e_obter_dados()
        if dados is None:
            return

        val, venc, status, f_id = dados

        # Valida duplicatas antes de gravar qualquer coisa
        duplicatas_validadas = []
        for i, d in enumerate(self.duplicatas, 1):
            v_str  = d["val_var"].get().strip()
            vc_str = d["venc_var"].get().strip()
            if not v_str or not vc_str:
                messagebox.showwarning("Aviso", f"Preencha valor e vencimento da Duplicata {i}.", parent=self)
                return
            try:
                v = float(v_str.replace(",", "."))
            except ValueError:
                messagebox.showerror("Erro", f"Valor invalido na Duplicata {i}.", parent=self)
                return
            try:
                datetime.strptime(vc_str, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Erro", f"Data invalida na Duplicata {i}. Use DD/MM/AAAA.", parent=self)
                return
            duplicatas_validadas.append({"valor": v, "vencimento": vc_str})

        conn = conectar()
        cursor = conn.cursor()

        if self.conta_id:
            cursor.execute("""
                UPDATE contas SET descricao=?, valor=?, vencimento=?, status=?, fornecedor_id=? WHERE id=?
            """, ("", val, venc, status, f_id, self.conta_id))
        else:
            cursor.execute("""
                INSERT INTO contas (descricao, valor, vencimento, status, fornecedor_id) VALUES (?, ?, ?, ?, ?)
            """, ("", val, venc, status, f_id))
            for dup in duplicatas_validadas:
                cursor.execute("""
                    INSERT INTO contas (descricao, valor, vencimento, status, fornecedor_id) VALUES (?, ?, ?, ?, ?)
                """, ("", dup["valor"], dup["vencimento"], status, f_id))

        conn.commit()
        conn.close()

        total = 1 + len(duplicatas_validadas)
        msg = f"{total} conta(s) salva(s) com sucesso!" if total > 1 else "Conta salva com sucesso!"
        messagebox.showinfo("Sucesso", msg, parent=self)
        self.callback_atualizar()

        if fechar:
            self.destroy()
        else:
            # Limpa tudo para proximo cadastro
            self.entry_valor.delete(0, tk.END)
            self.entry_vencimento.delete(0, tk.END)
            self.combo_fornecedor.set("(Nenhum)")
            self.combo_status.set("Pendente")
            # Remove todas as linhas de duplicata
            for d in self.duplicatas:
                d["num_label"].destroy()
                d["frame_linha"].destroy()
            self.duplicatas.clear()
            if not self.conta_id and hasattr(self, "sep_dup"):
                self.sep_dup.grid_remove()
            self._ajustar_tamanho()

    def salvar(self):
        self._gravar(fechar=False)

    def salvar_e_fechar(self):
        self._gravar(fechar=True)


# ============================================================
# JANELA: BOLETOS PAGOS
# ============================================================

class JanelaPagos(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Boletos Pagos")
        self.geometry("700x500")
        self.resizable(True, True)
        self.grab_set()
        self.criar_widgets()
        self.carregar_pagos()

    def criar_widgets(self):
        # --- Filtro por período ---
        frame_filtros = tk.LabelFrame(self, text="Filtro por Período de Pagamento", pady=5, padx=10)
        frame_filtros.pack(fill="x", padx=10, pady=10)

        tk.Label(frame_filtros, text="De:").pack(side="left", padx=2)
        self.entry_ini = DateEntry(frame_filtros, width=12, background="darkblue",
                                   foreground="white", borderwidth=2, date_pattern="dd/mm/yyyy",
                                   locale="pt_BR")
        self.entry_ini.pack(side="left", padx=5)

        tk.Label(frame_filtros, text="Até:").pack(side="left", padx=2)
        self.entry_fim = DateEntry(frame_filtros, width=12, background="darkblue",
                                   foreground="white", borderwidth=2, date_pattern="dd/mm/yyyy",
                                   locale="pt_BR")
        self.entry_fim.pack(side="left", padx=5)

        tk.Button(frame_filtros, text="🔍 Filtrar", command=self.carregar_pagos,
                  bg="#009688", fg="white", width=10).pack(side="left", padx=10)
        tk.Button(frame_filtros, text="📊 Exportar Excel", command=self.exportar_excel,
                  bg="#4CAF50", fg="white", width=15).pack(side="right", padx=5)

        # Define semana atual como padrão
        hoje = datetime.now().date()
        inicio_semana = hoje - __import__("datetime").timedelta(days=hoje.weekday())
        fim_semana = inicio_semana + __import__("datetime").timedelta(days=6)
        self.entry_ini.set_date(inicio_semana)
        self.entry_fim.set_date(fim_semana)

        # --- Tabela ---
        frame_grid = tk.Frame(self)
        frame_grid.pack(fill="both", expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(frame_grid,
                                  columns=("ID", "Fornecedor", "Valor", "Vencimento", "Pago Em"),
                                  show="headings", selectmode="browse")
        self.tree.heading("ID", text="ID")
        self.tree.heading("Fornecedor", text="Fornecedor")
        self.tree.heading("Valor", text="Valor")
        self.tree.heading("Vencimento", text="Vencimento")
        self.tree.heading("Pago Em", text="Pago Em")
        self.tree["displaycolumns"] = ("Fornecedor", "Valor", "Pago Em")
        self.tree.column("ID", width=0, stretch=False)
        self.tree.column("Fornecedor", width=300, anchor="center")
        self.tree.column("Valor", width=180, anchor="center")
        self.tree.column("Vencimento", width=0, stretch=False)
        self.tree.column("Pago Em", width=180, anchor="center")
        self.tree.tag_configure("pago", background="#C8F5C8", foreground="black")

        scrollbar = ttk.Scrollbar(frame_grid, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # --- Rodapé total ---
        frame_rodape = tk.Frame(self, bg="#f0f0f0", pady=8)
        frame_rodape.pack(fill="x", padx=10, pady=(0, 5))
        tk.Label(frame_rodape, text="Total Pago:", font=("Arial", 11, "bold"),
                 bg="#f0f0f0").pack(side="left", padx=10)
        self.label_total = tk.Label(frame_rodape, text="R$ 0,00",
                                     font=("Arial", 12, "bold"), fg="#27ae60", bg="#f0f0f0")
        self.label_total.pack(side="left")

    def carregar_pagos(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        data_ini = self.entry_ini.get().strip()
        data_fim = self.entry_fim.get().strip()

        query = """
            SELECT c.id, f.nome, c.valor, c.vencimento, c.data_pagamento
            FROM contas c LEFT JOIN fornecedores f ON c.fornecedor_id = f.id
            WHERE LOWER(TRIM(c.status)) = 'pago'
        """
        params = []

        if data_ini and data_fim:
            try:
                d_ini_iso = datetime.strptime(data_ini, "%d/%m/%Y").strftime("%Y-%m-%d")
                d_fim_iso = datetime.strptime(data_fim, "%d/%m/%Y").strftime("%Y-%m-%d")
                query += """ AND (
                    CASE WHEN c.data_pagamento IS NOT NULL AND c.data_pagamento != ''
                    THEN substr(c.data_pagamento,7,4)||'-'||substr(c.data_pagamento,4,2)||'-'||substr(c.data_pagamento,1,2)
                    ELSE substr(c.vencimento,7,4)||'-'||substr(c.vencimento,4,2)||'-'||substr(c.vencimento,1,2)
                    END
                ) BETWEEN ? AND ?"""
                params.extend([d_ini_iso, d_fim_iso])
            except ValueError:
                messagebox.showerror("Erro", "Formato de data inconsistente!", parent=self)
                return

        query += " ORDER BY substr(c.vencimento,7,4)||'-'||substr(c.vencimento,4,2)||'-'||substr(c.vencimento,1,2) ASC"

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute(query, params)
        self.dados = cursor.fetchall()
        conn.close()

        total = 0.0
        for linha in self.dados:
            c_id, forn_nome, valor, vencimento, data_pagamento = linha
            nome = forn_nome if forn_nome else "(Nenhum)"
            try:
                v_num = float(valor)
                valor_fmt = f"R$ {v_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except (ValueError, TypeError):
                v_num = 0.0
                valor_fmt = f"R$ {valor}"
            total += v_num
            pago_em = data_pagamento if data_pagamento else vencimento
            self.tree.insert("", tk.END, iid=c_id,
                              values=(c_id, nome, valor_fmt, vencimento, pago_em),
                              tags=("pago",))

        total_fmt = f"R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        self.label_total.config(text=total_fmt)

    def exportar_excel(self):
        if not EXCEL_DISPONIVEL:
            messagebox.showerror("Erro", "A biblioteca \'openpyxl\' não está instalada.", parent=self)
            return

        if not self.tree.get_children():
            messagebox.showwarning("Aviso", "Não há dados para exportar.", parent=self)
            return

        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar Relatório de Pagos",
            parent=self
        )
        if not arquivo:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boletos Pagos"
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = ws.page_margins.right = 0.5
        ws.page_margins.top = ws.page_margins.bottom = 0.5

        ws.append(["Fornecedor", "Valor (R$)", "Pago Em"])

        valor_total = 0.0
        for linha in self.dados:
            _, forn_nome, valor, vencimento, data_pagamento = linha
            nome = forn_nome if forn_nome else "(Nenhum)"
            try:
                v_num = float(valor)
            except:
                v_num = 0.0
            valor_total += v_num
            pago_em_str = data_pagamento if data_pagamento else vencimento
            try:
                pago_em_excel = datetime.strptime(pago_em_str, "%d/%m/%Y")
            except:
                pago_em_excel = pago_em_str
            ws.append([nome, v_num, pago_em_excel])

        ws.append([])
        linha_total_index = ws.max_row + 1
        ws.cell(row=linha_total_index, column=1, value="TOTAL PAGO:")
        ws.cell(row=linha_total_index, column=2, value=valor_total)

        # Estilização
        font_cab   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_dados = Font(name="Arial", size=10)
        font_total = Font(name="Arial", size=10, bold=True)
        fill_cab   = PatternFill(start_color="1A5C38", end_color="1A5C38", fill_type="solid")
        fill_total = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        align_c    = Alignment(horizontal="center", vertical="center")
        align_e    = Alignment(horizontal="left",   vertical="center")
        borda      = Border(
            left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF")
        )

        for cell in ws[1]:
            cell.font = fill_c = font_cab
            cell.fill = fill_cab
            cell.alignment = align_c
            cell.border = borda

        for row in ws.iter_rows(min_row=2, max_row=linha_total_index - 1, min_col=1, max_col=3):
            for cell in row:
                cell.font = font_dados
                cell.border = borda
                cell.alignment = align_e if cell.column == 1 else align_c
                if cell.column == 2:
                    cell.number_format = "\"R$\" #,##0.00"
                if cell.column == 3:
                    cell.number_format = "DD/MM/YYYY"

        for col in range(1, 4):
            cell = ws.cell(row=linha_total_index, column=col)
            cell.font = font_total
            cell.fill = fill_total
            cell.border = borda
            cell.alignment = align_e if col == 1 else align_c
            if col == 2:
                cell.number_format = "\"R$\" #,##0.00"

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 22

        try:
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", "Relatório gerado com sucesso!", parent=self)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar:\n{e}", parent=self)


# ============================================================
# JANELA PRINCIPAL (ROOT)
# ============================================================

class JanelaPrincipal:
    def __init__(self, root):
        self.root = root
        self.root.title("Controle Financeiro")
        self.root.geometry("700x450")
        self.criar_widgets()
        self.carregar_contas()

    def criar_widgets(self):
        # --- Bloco de Botões Superiores ---
        frame_botoes = tk.Frame(self.root, pady=10)
        frame_botoes.pack(fill="x", padx=10)

        tk.Button(frame_botoes, text="➕ Nova Conta", command=self.nova_conta, bg="#2196F3", fg="white", width=12).pack(
            side="left", padx=3)
        tk.Button(frame_botoes, text="✏️ Editar Conta", command=self.editar_conta, bg="#FF9800", fg="white",
                  width=12).pack(side="left", padx=3)
        tk.Button(frame_botoes, text="❌ Excluir Conta", command=self.excluir_conta, bg="#f44336", fg="white",
                  width=12).pack(side="left", padx=3)
        tk.Button(frame_botoes, text="👥 Fornecedores", command=self.abrir_fornecedores, bg="#9C27B0", fg="white",
                  width=12).pack(side="right", padx=3)
        tk.Button(frame_botoes, text="💰 Pagos", command=self.abrir_pagos, bg="#00796B", fg="white",
                  width=10).pack(side="right", padx=3)
        tk.Button(
            frame_botoes,
            text="✔ Dar Baixa",
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold"),
            command=self.dar_baixa
        ).pack(side="left", padx=5)

        # --- Bloco de Filtros por Período com Calendário Dinâmico ---
        frame_filtros = tk.LabelFrame(self.root, text="Filtro de Relatório por Vencimento", pady=5, padx=10)
        frame_filtros.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_filtros, text="De:").pack(side="left", padx=2)
        # Componente DateEntry: cria um campo de data formatado em PT-BR com calendário integrado
        self.entry_data_ini = DateEntry(frame_filtros, width=12, background='darkblue',
                                        foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                                        locale='pt_BR')
        self.entry_data_ini.pack(side="left", padx=5)

        tk.Label(frame_filtros, text="Até:").pack(side="left", padx=2)
        self.entry_data_fim = DateEntry(frame_filtros, width=12, background='darkblue',
                                        foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                                        locale='pt_BR')
        self.entry_data_fim.pack(side="left", padx=5)

        tk.Button(frame_filtros, text="🔍 Filtrar", command=self.filtrar_contas, bg="#009688", fg="white",
                  width=10).pack(side="left", padx=10)
        tk.Button(frame_filtros, text="📊 Exportar Excel", command=self.exportar_excel, bg="#4CAF50", fg="white",
                  width=15).pack(side="right", padx=5)

        # --- Grid da Tabela ---
        frame_grid = tk.Frame(self.root)
        frame_grid.pack(fill="both", expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(frame_grid, columns=("ID", "Fornecedor", "Valor", "Vencimento", "Status"), show="headings", selectmode="extended")
        self.tree.heading("ID", text="ID")
        self.tree.heading("Fornecedor", text="Fornecedor")
        self.tree.heading("Valor", text="Valor")
        self.tree.heading("Vencimento", text="Vencimento")
        self.tree.heading("Status", text="Status")

        self.tree["displaycolumns"] = (
    "Fornecedor",
    "Valor",
    "Vencimento",
    "Status"
)

        self.tree.column("ID", width=0, stretch=False)
        self.tree.column("Fornecedor", width=250, anchor="center")
        self.tree.column("Valor", width=120, anchor="center")
        self.tree.column("Vencimento", width=120, anchor="center")
        self.tree.column("Status", width=100, anchor="center")
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure('vencido', background='#FFCCCC', foreground='black')
        self.tree.tag_configure('pago', background='#C8F5C8', foreground='black')

        # --- Rodapé com Total a Pagar ---
        frame_rodape = tk.Frame(self.root, bg="#f0f0f0", pady=8)
        frame_rodape.pack(fill="x", padx=10, pady=(0, 5))

        tk.Label(frame_rodape, text="Total a Pagar (Pendentes):", font=("Arial", 11, "bold"),
                 bg="#f0f0f0").pack(side="left", padx=10)
        self.label_total = tk.Label(frame_rodape, text="R$ 0,00", font=("Arial", 12, "bold"),
                                    fg="#c0392b", bg="#f0f0f0")
        self.label_total.pack(side="left")

    def dar_baixa(self):

        itens_selecionados = self.tree.selection()

        if not itens_selecionados:
            messagebox.showwarning(
                "Aviso",
                "Selecione ao menos uma conta."
            )
            return

        resposta = messagebox.askyesno(
            "Confirmar",
            f"Deseja dar baixa em {len(itens_selecionados)} boleto(s)?"
        )

        if not resposta:
            return

        try:

            conn = conectar()
            cursor = conn.cursor()

            for item_id in itens_selecionados:
                item = self.tree.item(item_id)

                valores = item['values']
                print(valores)

                id_conta = valores[0]

                status = valores[3]

                if status == "Pago":
                    continue

                data_hoje = datetime.now().strftime("%d/%m/%Y")

                cursor.execute("""
                    UPDATE contas
                    SET status = 'Pago', data_pagamento = ?
                    WHERE id = ?
                """, (data_hoje, id_conta,))

            conn.commit()
            conn.close()

            messagebox.showinfo(
                "Sucesso",
                "Baixa realizada com sucesso!"
            )

            self.carregar_contas()

        except Exception as e:
            messagebox.showerror(
                "Erro",
                f"Erro ao atualizar status:\n{e}"
            )

    def carregar_contas(self):
        """Carga padrão ao iniciar: mostra todos os boletos pendentes e vencidos."""
        query = """
            SELECT c.id, f.nome, c.valor, c.vencimento, c.status 
            FROM contas c LEFT JOIN fornecedores f ON c.fornecedor_id = f.id
            WHERE LOWER(TRIM(c.status)) != 'pago'
            ORDER BY substr(c.vencimento,7,4)||'-'||substr(c.vencimento,4,2)||'-'||substr(c.vencimento,1,2) ASC
        """
        self._executar_query(query, [])

    def filtrar_contas(self):
        """Filtro acionado pelo botão: filtra pelo período de vencimento selecionado."""
        data_ini = self.entry_data_ini.get().strip()
        data_fim = self.entry_data_fim.get().strip()

        if not data_ini or not data_fim:
            messagebox.showwarning("Aviso", "Selecione as datas de início e fim para filtrar.")
            return

        try:
            d_ini_iso = datetime.strptime(data_ini, "%d/%m/%Y").strftime("%Y-%m-%d")
            d_fim_iso = datetime.strptime(data_fim, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Erro", "Formato interno de data inconsistente!")
            return

        query = """
            SELECT c.id, f.nome, c.valor, c.vencimento, c.status 
            FROM contas c LEFT JOIN fornecedores f ON c.fornecedor_id = f.id
            WHERE substr(c.vencimento,7,4)||'-'||substr(c.vencimento,4,2)||'-'||substr(c.vencimento,1,2) BETWEEN ? AND ?
            ORDER BY substr(c.vencimento,7,4)||'-'||substr(c.vencimento,4,2)||'-'||substr(c.vencimento,1,2) ASC
        """
        self._executar_query(query, [d_ini_iso, d_fim_iso])

    def _executar_query(self, query, params):
        """Executa a query e popula a tabela com os resultados."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        data_hoje_obj = datetime.now().date()

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute(query, params)
        self.dados_carregados = cursor.fetchall()
        conn.close()

        for linha in self.dados_carregados:
            c_id, forn_nome, valor, vencimento, status = linha

            nome_fornecedor = forn_nome if forn_nome else "(Nenhum)"

            try:
                valor_formatado = f"R$ {float(valor):.2f}"
            except (ValueError, TypeError):
                valor_formatado = f"R$ {valor}"

            # --- Lógica de Validação de Vencimento para as Tags ---
            tags_linha = ()
            try:
                data_venc_obj = datetime.strptime(vencimento, "%d/%m/%Y").date()
                if str(status).strip().lower() == "pago":
                    tags_linha = ('pago',)
                elif data_venc_obj < data_hoje_obj:
                    tags_linha = ('vencido',)
            except (ValueError, TypeError):
                pass

            self.tree.insert(
                "",
                tk.END,
                iid=c_id,
                values=(
                    c_id,
                    nome_fornecedor,
                    valor_formatado,
                    vencimento,
                    status
                ),
                tags=tags_linha
            )

        # Atualiza o total sempre com o valor real do banco (pendentes + vencidos)
        self._atualizar_total()


    def _atualizar_total(self):
        """Busca direto do banco o total de todos os boletos pendentes/vencidos, ignorando filtros."""
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(SUM(valor), 0)
            FROM contas
            WHERE LOWER(TRIM(status)) != 'pago'
        """)
        total = cursor.fetchone()[0]
        conn.close()
        self.label_total.config(
            text=f"R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )

    def exportar_excel(self):
        """Exporta os dados diretamente do Banco de Dados para garantir o alinhamento perfeito."""
        if not EXCEL_DISPONIVEL:
            messagebox.showerror("Erro", "A biblioteca 'openpyxl' não está instalada.")
            return

        if not self.tree.get_children():
            messagebox.showwarning("Aviso", "Não há dados na tabela para exportar.")
            return

        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar Relatório Financeiro"
        )

        if not arquivo:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Financeiro"

        # 1. Configuração de Página para Impressão A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # 2. Escreve o Cabeçalho
        headers = ["Fornecedor", "Valor (R$)", "Vencimento", "Status"]
        ws.append(headers)

        # 3. Busca os dados direto do Banco
        data_ini = self.entry_data_ini.get().strip()
        data_fim = self.entry_data_fim.get().strip()

        query = """
            SELECT f.nome, c.valor, c.vencimento, c.status 
            FROM contas c LEFT JOIN fornecedores f ON c.fornecedor_id = f.id
        """
        params = []

        if data_ini and data_fim:
            try:
                d_ini_iso = datetime.strptime(data_ini, "%d/%m/%Y").strftime("%Y-%m-%d")
                d_fim_iso = datetime.strptime(data_fim, "%d/%m/%Y").strftime("%Y-%m-%d")
                query += " WHERE substr(c.vencimento,7,4)||'-'||substr(c.vencimento,4,2)||'-'||substr(c.vencimento,1,2) BETWEEN ? AND ?"
                params.extend([d_ini_iso, d_fim_iso])
            except ValueError:
                pass

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute(query, params)
        dados_banco = cursor.fetchall()
        conn.close()

        # 4. Preenche as linhas do Excel com os dados limpos do banco
        valor_total = 0.0
        for linha in dados_banco:
            forn_nome, valor, vencimento, status = linha
            nome_fornecedor = forn_nome if forn_nome else "(Nenhum)"

            try:
                v_num = float(valor)
            except:
                v_num = 0.0

            valor_total += v_num

            try:
                data_excel = datetime.strptime(vencimento, "%d/%m/%Y")
            except:
                data_excel = vencimento

            ws.append([nome_fornecedor, v_num, data_excel, status])

        # Linha em branco
        ws.append([])

        # Linha total
        linha_total_index = ws.max_row + 1

        ws.cell(row=linha_total_index, column=1, value="TOTAL GERAL:")
        ws.cell(row=linha_total_index, column=2, value=valor_total)

        # 6. Estilização Visual Unificada e Centralização Total das Contas
        font_cabecalho = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_dados = Font(name="Arial", size=10, bold=False)
        font_total = Font(name="Arial", size=10, bold=True)

        fill_cabecalho = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        align_centro = Alignment(horizontal="center", vertical="center")
        align_esquerda = Alignment(horizontal="left", vertical="center")

        borda_fina = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )

        # Formata o Cabeçalho (Linha 1)
        # Cabeçalho
        for cell in ws[1]:
            cell.font = font_cabecalho
            cell.fill = fill_cabecalho
            cell.alignment = align_centro
            cell.border = borda_fina

        # Dados
        for row in ws.iter_rows(
                min_row=2,
                max_row=linha_total_index - 1,
                min_col=1,
                max_col=4
        ):
            for cell in row:
                cell.font = font_dados
                cell.border = borda_fina

                if cell.column == 1:
                    cell.alignment = align_esquerda
                else:
                    cell.alignment = align_centro

                if cell.column == 2:
                    cell.number_format = '"R$" #,##0.00'

                if cell.column == 3:
                    cell.number_format = 'DD/MM/YYYY'

        # Total
        for col in range(1, 5):
            cell = ws.cell(row=linha_total_index, column=col)

            cell.font = font_total
            cell.fill = fill_total
            cell.border = borda_fina

            if col == 1:
                cell.alignment = align_esquerda
            else:
                cell.alignment = align_centro

            if col == 2:
                cell.number_format = '"R$" #,##0.00'

        # Largura fixa
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        # Altura padrão
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 22

        # 7. Salva o arquivo e exibe a mensagem de sucesso
        try:
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", "Relatório gerado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o arquivo:\n{e}")

    def nova_conta(self):
        JanelaConta(self.root, self.carregar_contas)

    def editar_conta(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione uma conta para editar.")
            return
        conta_id = self.tree.item(sel[0])['values'][0]
        JanelaConta(self.root, self.carregar_contas, conta_id)

    def excluir_conta(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione uma conta para excluir.")
            return

        # Correção: Pega o primeiro item selecionado da tupla do Treeview
        conta_id = sel[0]

        if messagebox.askyesno("Confirmar", "Deseja realmente excluir esta conta?"):
            conn = conectar()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM contas WHERE id = ?", (conta_id,))
            conn.commit()
            conn.close()

            # Recarrega a tabela na tela atualizando os dados
            self.carregar_contas()
            messagebox.showinfo("Sucesso", "Conta excluída com sucesso!")

    def abrir_fornecedores(self):
        JanelaFornecedores(self.root, self.carregar_contas)

    def abrir_pagos(self):
        JanelaPagos(self.root)


if __name__ == "__main__":
    criar_tabelas()
    root = tk.Tk()
    app = JanelaPrincipal(root)
    root.mainloop()
